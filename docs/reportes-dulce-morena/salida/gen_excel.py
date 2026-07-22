# Genera el Excel del reporte de ventas dulce morena:
#  - Hoja 1: Top 10 por unidades por categoría (último año)
#  - Hoja 2: Top 10 por facturación por categoría (último año)
#  - Hoja 3: Comparativo trimestral 2024-2026 + proyección T3/T4-2026
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP = Path(__file__).parent
OUT = SP / "reporte_ventas_dulce_morena.xlsx"

AZUL = "1F4E79"
GRIS = "D9E2F3"
AMARILLO = "FFF2CC"   # columnas proyectadas
BORDE = Border(bottom=Side(style="thin", color="BFBFBF"))
H_FONT = Font(bold=True, color="FFFFFF")
H_FILL = PatternFill("solid", fgColor=AZUL)
CAT_FONT = Font(bold=True, color=AZUL)
CAT_FILL = PatternFill("solid", fgColor=GRIS)
PROY_FILL = PatternFill("solid", fgColor=AMARILLO)
NOTA_FONT = Font(italic=True, size=9, color="595959")

def leer_tsv(nombre):
    with open(SP / nombre, encoding="utf-8") as f:
        filas = list(csv.reader(f, delimiter="\t"))
    return filas[0], [r for r in filas[1:] if r]

def hoja_top(ws, tsv, titulo, orden_col):
    ws.sheet_view.showGridLines = False
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws["A2"] = ("Período: 18-jul-2025 al 18-jul-2026 · Cajas 1, 3 y 4 consolidadas · "
                "Excluye facturas anuladas y categoría TECNO")
    ws["A2"].font = NOTA_FONT

    _, datos = leer_tsv(tsv)
    headers = ["#", "Producto", "Unidades", "Total (L)", "% de la categoría"]
    fila = 4
    cat_actual = None
    for r in datos:
        cat = r[0]
        if cat != cat_actual:
            cat_actual = cat
            total_cat = float(r[5]) if orden_col == "unidades" else None
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
            c = ws.cell(fila, 1, f"{cat}" + (f"  —  L {total_cat:,.0f} en el año" if total_cat else ""))
            c.font = CAT_FONT
            for col in range(1, 6):
                ws.cell(fila, col).fill = CAT_FILL
            fila += 1
            for col, h in enumerate(headers, 1):
                c = ws.cell(fila, col, h)
                c.font = H_FONT
                c.fill = H_FILL
                c.alignment = Alignment(horizontal="center")
            fila += 1
        ws.cell(fila, 1, int(r[1])).alignment = Alignment(horizontal="center")
        ws.cell(fila, 2, r[2])
        ws.cell(fila, 3, float(r[3])).number_format = "#,##0"
        ws.cell(fila, 4, float(r[4])).number_format = "#,##0.00"
        if orden_col == "lps":
            ws.cell(fila, 5, float(r[5]) / 100).number_format = "0.0%"
        else:
            ws.cell(fila, 5, float(r[4]) / float(r[5])).number_format = "0.0%"
        for col in range(1, 6):
            ws.cell(fila, col).border = BORDE
        fila += 1
    for i, a in enumerate([6, 48, 12, 14, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = "A4"

def hoja_trimestres(ws):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Comparativo trimestral 2024–2026 y proyección de cierre 2026"
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws["A2"] = ("Cajas 1, 3 y 4 consolidadas · Excluye anuladas y TECNO · T3-2026 real solo del 1 al 18-jul · "
                "Columnas amarillas = proyección · Cafetería abrió en T3-2025")
    ws["A2"].font = NOTA_FONT

    _, datos = leer_tsv("trimestres_full.tsv")
    tabla = {}   # {cat: {(anio,tri): (unid, lps)}}
    for anio, tri, cat, unid, lps in datos:
        tabla.setdefault(cat, {})[(int(anio), int(tri))] = (float(unid), float(lps))

    _, jul = leer_tsv("julio_ventana.tsv")
    ventana = {}  # {cat: {anio: (unid, lps)}}
    for anio, cat, unid, lps in jul:
        ventana.setdefault(cat, {})[int(anio)] = (float(unid), float(lps))

    cats = ["Pasteleria", "Panaderia", "Reposteria", "Cafeteria"]
    periodos = [(2024, t) for t in (1, 2, 3, 4)] + [(2025, t) for t in (1, 2, 3, 4)] + [(2026, 1), (2026, 2), (2026, 3)]

    # Factor de proyección por categoría (índice 0=unidades, 1=L):
    # ventas 1-ene→18-jul 2026 / mismo período 2025 (T1+T2+ventana julio)
    def factor(cat, idx):
        v25, v26 = ventana.get(cat, {}).get(2025), ventana.get(cat, {}).get(2026)
        t = tabla.get(cat, {})
        if not (v25 and v26 and (2025, 1) in t and (2025, 2) in t):
            return None
        ytd25 = t[(2025, 1)][idx] + t[(2025, 2)][idx] + v25[idx]
        ytd26 = t[(2026, 1)][idx] + t[(2026, 2)][idx] + v26[idx]
        return ytd26 / ytd25

    def proyeccion(cat, tri, idx):
        t = tabla.get(cat, {})
        f = factor(cat, idx)
        if f is not None and (2025, tri) in t:
            return t[(2025, tri)][idx] * f
        # Cafetería: sin base 2025 comparable
        if tri == 3 and (2026, 3) in t:
            return t[(2026, 3)][idx] * 92 / 18          # ritmo diario del parcial
        if tri == 4 and (2025, 4) in t:
            return t[(2025, 4)][idx]                     # se asume repetir temporada
        return None

    def bloque(fila0, titulo, idx, fmt):
        ws.cell(fila0, 1, titulo).font = Font(bold=True, size=12, color=AZUL)
        fila = fila0 + 1
        heads = (["Categoría"] + [f"T{t}-{a % 100}" for a, t in periodos[:8]]
                 + ["T1-26", "T2-26", "T3-26 *", "T3-26 py", "T4-26 py",
                    "Año 2024", "Año 2025", "2026 py", "Var 25/24", "Var 26py/25"])
        for col, h in enumerate(heads, 1):
            c = ws.cell(fila, col, h)
            c.font = H_FONT
            c.fill = H_FILL
            c.alignment = Alignment(horizontal="center")
        fila += 1
        col_proy = {13, 14, 17}          # T3 py, T4 py, 2026 py
        tot = [0.0] * 20
        for cat in cats:
            t = tabla.get(cat, {})
            ws.cell(fila, 1, cat)
            vals = []
            for p in periodos:
                vals.append(t.get(p, (None, None))[idx] if p in t else None)
            p3, p4 = proyeccion(cat, 3, idx), proyeccion(cat, 4, idx)
            a24 = sum(t[p][idx] for p in t if p[0] == 2024) or None
            a25 = sum(t[p][idx] for p in t if p[0] == 2025) or None
            a26 = (t.get((2026, 1), (0, 0))[idx] + t.get((2026, 2), (0, 0))[idx]
                   + (p3 or 0) + (p4 or 0)) or None
            fila_vals = vals + [p3, p4, a24, a25, a26]
            for col, v in enumerate(fila_vals, 2):
                if v is None:
                    ws.cell(fila, col, "—").alignment = Alignment(horizontal="center")
                else:
                    ws.cell(fila, col, v).number_format = fmt
                    tot[col] += v
                if col in col_proy:
                    ws.cell(fila, col).fill = PROY_FILL
            # variaciones anuales (Cafetería: bases incompletas → —)
            var2524 = (a25 / a24 - 1) if (a24 and a25 and cat != "Cafeteria") else None
            var2625 = (a26 / a25 - 1) if (a25 and a26 and cat != "Cafeteria") else None
            for col, v in [(18, var2524), (19, var2625)]:
                if v is None:
                    ws.cell(fila, col, "—").alignment = Alignment(horizontal="center")
                else:
                    ws.cell(fila, col, v).number_format = "+0.0%;-0.0%"
            for col in range(1, 20):
                ws.cell(fila, col).border = BORDE
            fila += 1
        ws.cell(fila, 1, "TOTAL").font = Font(bold=True)
        ws.cell(fila, 1).fill = CAT_FILL
        for col in range(2, 18):
            c = ws.cell(fila, col, tot[col])
            c.number_format = fmt
            c.font = Font(bold=True)
            c.fill = PROY_FILL if col in col_proy else CAT_FILL
        for col, (na, nb) in [(18, (15, 16)), (19, (16, 17))]:  # Var sobre años totales
            c = ws.cell(fila, col, tot[nb] / tot[na] - 1)
            c.number_format = "+0.0%;-0.0%"
            c.font = Font(bold=True)
            c.fill = CAT_FILL
        return fila + 2

    fila = bloque(4, "Facturación (L)", 1, "#,##0")
    fila = bloque(fila, "Unidades vendidas", 0, "#,##0")
    notas = [
        "* T3-2026 real: solo ventas del 1 al 18 de julio de 2026.",
        "py = proyección. Método: T3/T4-2026 = T3/T4-2025 × factor de crecimiento de la categoría, donde el factor",
        "compara las ventas del 1-ene al 18-jul de 2026 contra el mismo período de 2025 (alineado estacionalmente).",
        "Cafetería (sin base 2025 comparable): T3 = ritmo diario del parcial × 92 días; T4 = se asume igual a T4-2025.",
        "2026 py = T1 + T2 reales + T3 py + T4 py. Var 26py/25 compara esa proyección contra 2025 real.",
        "Cafetería sin variaciones anuales: 2025 solo tiene medio año (abrió en T3-2025).",
    ]
    for n in notas:
        ws.cell(fila, 1, n).font = NOTA_FONT
        fila += 1
    ws.column_dimensions["A"].width = 13
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 10
    for col in range(15, 20):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.freeze_panes = "B6"

wb = Workbook()
hoja_top(wb.active, "top_unidades.tsv", "Top 10 por unidades vendidas", "unidades")
wb.active.title = "Top 10 unidades"
ws2 = wb.create_sheet("Top 10 facturación")
hoja_top(ws2, "top_lps.tsv", "Top 10 por facturación (L)", "lps")
ws3 = wb.create_sheet("Trimestral 2024-2026 + proy")
hoja_trimestres(ws3)
wb.save(OUT)
print(f"OK: {OUT}")
