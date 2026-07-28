#!/usr/bin/env python3
"""
US-051 — Genera las plantillas Excel de carga inicial (productos y clientes)
para que el cliente (Urbina) llene su catálogo. Cada plantilla trae:
  - hoja de Instrucciones,
  - hoja de datos con encabezados, filas de ejemplo y validaciones (dropdowns),
  - hoja oculta "Listas" con los valores válidos (categorías, impuestos).

Las columnas se alinean con el contrato del sistema:
  ProductRequest{name, price, categoryId(=categoría), taxId(=impuesto), altCode}
  + Precio de costo → precios_articulos nivel codigo_precio=1 (costo), la misma
    fuente que leen ArticuloDao y el trigger de kardex para el margen.
  CustomerCreateRequest{name, rtn, address, phone}

Listas (categorías/impuestos) = snapshot de admin_tools al 2026-06-01.
Regenerar:  python3 docs/plantillas/generar_plantillas.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# El ISV es estándar (no depende del cliente): exonerado / 15% / 18%.
IMPUESTOS = ["Exectos (0%)", "Basicos (15%)", "Lujo (18%)"]

# El cliente define SUS propias categorías en la hoja "Categorias" (no imponemos
# las nuestras). Estos son solo ejemplos a reemplazar.
EJEMPLO_CATEGORIAS = ["ABARROTES", "BEBIDAS", "LIMPIEZA", "CUIDADO PERSONAL", "ROPA"]
MAX_CATEGORIAS = 200  # rango del dropdown que apunta a la hoja Categorias

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
REQ_FONT = Font(bold=True, color="FFFFFF", size=11)
EXAMPLE_FONT = Font(italic=True, color="888888")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def estilizar_encabezado(ws, ncols, fila=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=fila, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[fila].height = 28
    ws.freeze_panes = ws.cell(row=fila + 1, column=1)


def hoja_listas(wb, impuestos):
    """Hoja oculta con los valores fijos del sistema (solo ISV)."""
    ws = wb.create_sheet("Listas")
    ws["A1"] = "Impuestos"
    for i, v in enumerate(impuestos, start=2):
        ws.cell(row=i, column=1, value=v)
    ws.sheet_state = "hidden"
    return ws


def hoja_categorias(wb):
    """Hoja visible donde el cliente ingresa SUS categorías (paso 1)."""
    ws = wb.create_sheet("Categorias")
    ws.cell(row=1, column=1, value="Categoria *")
    ws.column_dimensions["A"].width = 30
    estilizar_encabezado(ws, 1)
    # Ejemplos a reemplazar por las categorías reales del cliente.
    for i, v in enumerate(EJEMPLO_CATEGORIAS, start=2):
        c = ws.cell(row=i, column=1, value=v)
        c.font = EXAMPLE_FONT
        c.border = BORDER
    return ws


def hoja_instrucciones(wb, titulo, lineas):
    ws = wb.create_sheet("Instrucciones", 0)
    ws["A1"] = titulo
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 110
    r = 3
    for ln in lineas:
        c = ws.cell(row=r, column=1, value=ln)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if ln.endswith(":") or ln.startswith("•") is False and ln.isupper():
            c.font = Font(bold=True)
        r += 1
    return ws


def dv(formula, prompt, allow_blank=True):
    d = DataValidation(type="list", formula1=formula, allow_blank=allow_blank, showDropDown=False)
    d.error = "Elegí un valor de la lista."
    d.errorTitle = "Valor inválido"
    d.prompt = prompt
    d.promptTitle = "Ayuda"
    return d


# ============================ PRODUCTOS ============================
def plantilla_productos():
    wb = Workbook()
    wb.remove(wb.active)
    hoja_listas(wb, IMPUESTOS)
    hoja_categorias(wb)  # paso 1: el cliente ingresa sus categorías

    ws = wb.create_sheet("Productos")
    headers = [
        ("Nombre *", 34),
        ("Categoria *", 20),
        ("Impuesto (ISV) *", 16),
        ("Precio de venta *", 16),
        ("Precio de costo", 16),
        ("Codigo de barras", 20),
        ("Codigo alterno", 16),
    ]
    for i, (h, w) in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    estilizar_encabezado(ws, len(headers))

    ejemplos = [
        ["Coca Cola 600ml", "BEBIDAS", "Basicos (15%)", 25.00, 18.00, "7501055300013", 1001],
        ["Camiseta algodon M", "ROPA", "Basicos (15%)", 180.00, 110.00, "", 1002],
        ["Jabon de tocador", "CUIDADO PERSONAL", "Basicos (15%)", 18.50, "", "", ""],
    ]
    for r, fila in enumerate(ejemplos, start=2):
        for c, val in enumerate(fila, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = BORDER

    n = 600
    last = 1 + len(ejemplos) + n
    # Categoría: dropdown que apunta a la hoja "Categorias" (lo que el cliente cargó).
    dv_cat = dv("=Categorias!$A$2:$A$%d" % (1 + MAX_CATEGORIAS),
                "Elegí una categoría de las que cargaste en la hoja 'Categorias'.")
    dv_imp = dv("=Listas!$A$2:$A$%d" % (1 + len(IMPUESTOS)), "Seleccioná el ISV que aplica.")
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_imp)
    dv_cat.add("B2:B%d" % last)
    dv_imp.add("C2:C%d" % last)

    dv_precio = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv_precio.error = "El precio debe ser un número mayor o igual a 0."
    dv_precio.errorTitle = "Precio inválido"
    ws.add_data_validation(dv_precio)
    dv_precio.add("D2:D%d" % last)

    # Precio de costo (columna E): misma regla numérica que el de venta.
    dv_costo = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv_costo.error = "El costo debe ser un número mayor o igual a 0 (o dejarse vacío)."
    dv_costo.errorTitle = "Costo inválido"
    ws.add_data_validation(dv_costo)
    dv_costo.add("E2:E%d" % last)

    instr = [
        "PLANTILLA DE CARGA INICIAL — PRODUCTOS",
        "",
        "IMPORTANTE: este archivo tiene DOS pasos en este orden:",
        "",
        "PASO 1 — Hoja \"Categorias\":",
        "• Ingresá PRIMERO todas TUS categorías (una por fila, en la columna 'Categoria').",
        "• Usá tus propios nombres de categoría — no hay una lista fija del sistema.",
        "• Las filas en gris son ejemplos: borralas y poné las tuyas.",
        "",
        "PASO 2 — Hoja \"Productos\" (una fila por producto):",
        "• Nombre *  (obligatorio): nombre del producto tal como aparecerá en facturas y catálogo.",
        "• Categoria *  (obligatorio): elegí del desplegable — solo aparecen las categorías que cargaste en el PASO 1.",
        "• Impuesto (ISV) *  (obligatorio): Exectos (0%) = exonerado, Basicos (15%), Lujo (18%). Elegí del desplegable.",
        "• Precio de venta *  (obligatorio): precio al público, en Lempiras, CON impuesto incluido. Solo números (ej. 25.00).",
        "• Precio de costo  (opcional): lo que te cuesta el producto (precio de compra), en Lempiras. Solo números (ej. 18.00). Se usa para calcular el margen y en los reportes de inventario; si no lo manejás, dejalo vacío.",
        "• Codigo de barras  (opcional): si el producto tiene código de barras escaneable.",
        "• Codigo alterno  (opcional): código interno/SKU si manejás uno.",
        "",
        "REGLAS:",
        "• Cargá las categorías ANTES de los productos (el desplegable de Categoría se llena desde la hoja 'Categorias').",
        "• No cambies los encabezados ni el orden de las columnas.",
        "• No dejes filas vacías en medio de los datos.",
        "• El precio incluye el ISV (el sistema calcula el desglose automáticamente).",
        "",
        "Cuando termines, guardá el archivo y envialo para la carga. Soporta varios cientos de productos por archivo.",
    ]
    hoja_instrucciones(wb, "Instrucciones — Productos", instr)

    path = os.path.join(OUT_DIR, "plantilla_productos.xlsx")
    wb.save(path)
    return path


# ============================ CLIENTES ============================
def plantilla_clientes():
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Clientes")
    headers = [
        ("Nombre *", 34),
        ("RTN", 20),
        ("Direccion", 40),
        ("Telefono", 18),
    ]
    for i, (h, w) in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    estilizar_encabezado(ws, len(headers))

    ejemplos = [
        ["Consumidor Final", "", "", ""],
        ["Distribuidora La Económica", "08011985123456", "Bo. El Centro, La Ceiba", "2443-1122"],
        ["María José Mayorga", "", "Col. Las Flores, calle 3", "9988-7766"],
    ]
    for r, fila in enumerate(ejemplos, start=2):
        for c, val in enumerate(fila, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = BORDER
            if c == 2:  # RTN como texto para no perder ceros
                cell.number_format = "@"

    n = 600
    last = 1 + len(ejemplos) + n
    # RTN: 14 dígitos exactos (o vacío). Validación por longitud de texto.
    dv_rtn = DataValidation(type="textLength", operator="equal", formula1="14", allow_blank=True)
    dv_rtn.error = "El RTN hondureño debe tener exactamente 14 dígitos (o dejarlo vacío)."
    dv_rtn.errorTitle = "RTN inválido"
    dv_rtn.prompt = "14 dígitos, sin guiones. Dejalo vacío para Consumidor Final."
    dv_rtn.promptTitle = "RTN"
    ws.add_data_validation(dv_rtn)
    dv_rtn.add("B2:B%d" % last)
    for r in range(2, last + 1):
        ws.cell(row=r, column=2).number_format = "@"

    instr = [
        "PLANTILLA DE CARGA INICIAL — CLIENTES",
        "",
        "Llená la hoja \"Clientes\" (una fila por cliente). Las filas en gris claro son EJEMPLOS: borralas o reemplazalas.",
        "",
        "COLUMNAS:",
        "• Nombre *  (obligatorio): nombre o razón social del cliente.",
        "• RTN  (opcional): 14 dígitos, sin guiones. Dejalo vacío para 'Consumidor Final'. (validación de 14 dígitos activada)",
        "• Direccion  (opcional).",
        "• Telefono  (opcional).",
        "",
        "REGLAS:",
        "• No cambies los encabezados ni el orden de las columnas.",
        "• No dejes filas vacías en medio de los datos.",
        "• El RTN debe tener 14 dígitos exactos o quedar vacío.",
        "",
        "Cuando termines, guardá el archivo y envialo para la carga.",
    ]
    hoja_instrucciones(wb, "Instrucciones — Clientes", instr)

    path = os.path.join(OUT_DIR, "plantilla_clientes.xlsx")
    wb.save(path)
    return path


if __name__ == "__main__":
    print("Generada:", plantilla_productos())
    print("Generada:", plantilla_clientes())
